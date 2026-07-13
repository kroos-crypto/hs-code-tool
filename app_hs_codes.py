import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from hs_code_classifier import classify_hs_code, get_notes
import io

st.set_page_config(page_title="HS Code Filler", layout="wide")
st.title("HS Code Auto-Filler")
st.markdown("Upload your Excel file – HS Codes will be automatically assigned")

uploaded_file = st.file_uploader("Upload Excel file here", type=['xlsx', 'xls'])

if uploaded_file:
    st.success("✓ File uploaded successfully!")
    df = pd.read_excel(uploaded_file, sheet_name='Data sheet')
    unique = df.drop_duplicates(subset=['Style number']).copy()
    st.info(str(len(unique)) + " unique product types found")

    results = []
    for idx, row in unique.iterrows():
        code = classify_hs_code(row)
        count = len(df[df['Style number'] == row['Style number']])
        results.append({
            'Product': row['Style/Display name'],
            'Category': row['Boozt Product Category'],
            'Gender': row.get('Gender (F = Female, M = Male, U = Unisex)', ''),
            'Material': row.get('Material composition', ''),
            'HS Code': code,
            'Variants': count,
            'Description': get_notes(code)
        })

    st.subheader("Automatic HS Code Classification")
    st.dataframe(pd.DataFrame(results), use_container_width=True, hide_index=True)

    # Build mapping with string keys to avoid int/str mismatch with openpyxl
    hs_mapping = {str(row['Style number']): classify_hs_code(row) for _, row in unique.iterrows()}

    if st.button("✨ Apply HS Codes & Download Excel", use_container_width=True):
        input_buffer = io.BytesIO(uploaded_file.getvalue())
        wb = load_workbook(input_buffer)
        ws = wb['Data sheet']

        col_idx = None
        style_col_idx = None
        for cell in ws[1]:
            if cell.value == 'Customs code/Tariff Code (Sweden)':
                col_idx = cell.column
            if cell.value == 'Style number':
                style_col_idx = cell.column

        if col_idx and style_col_idx:
            for row_idx in range(2, ws.max_row + 1):
                style_val = str(ws.cell(row=row_idx, column=style_col_idx).value)
                if style_val in hs_mapping:
                    ws.cell(row=row_idx, column=col_idx, value=hs_mapping[style_val])

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        st.success("✓ HS Codes successfully applied!")
        st.download_button(
            label="📥 Download Excel with HS Codes",
            data=output_buffer.getvalue(),
            file_name=Path(uploaded_file.name).stem + "_with_hs_codes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
else:
    st.info("👆 Please upload your Excel file to get started")
